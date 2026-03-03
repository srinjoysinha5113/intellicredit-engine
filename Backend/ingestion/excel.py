"""
Excel file ingestion.
Handles both .xlsx and .xls formats.
"""

from typing import List
from openpyxl import load_workbook
import pandas as pd

from config import CHUNK_SIZE, CHUNK_OVERLAP
from ingestion.chunking import split_text_into_chunks
from models.schemas import KnowledgeChunk


def process_excel_file(file_path: str, file: str, relative_path: str, category: str, subcategory: str) -> List[KnowledgeChunk]:
    """Process Excel file and create KnowledgeChunk objects."""
    chunks = []
    
    try:
        workbook = load_workbook(file_path, read_only=True)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_text = []
            
            for row in sheet.iter_rows(values_only=True):
                row_text = " ".join([str(cell) for cell in row if cell is not None])
                if row_text.strip():
                    sheet_text.append(row_text)
            
            if sheet_text:
                full_sheet_text = f"Sheet: {sheet_name}\n" + "\n".join(sheet_text)
                sheet_chunks = split_text_into_chunks(full_sheet_text, chunk_size=CHUNK_SIZE, overlap=CHUNK_OVERLAP)
                
                for chunk in sheet_chunks:
                    knowledge_chunk = KnowledgeChunk(
                        content=chunk,
                        source_type="excel",
                        source_name=file,
                        source_path=relative_path,
                        metadata={"category": category, "subcategory": subcategory, "sheet": sheet_name},
                        structure={"sheet": sheet_name}
                    )
                    chunks.append(knowledge_chunk)
        
        workbook.close()
        
    except Exception as e:
        print(f"Error processing Excel file {file}: {str(e)}")
    
    return chunks
